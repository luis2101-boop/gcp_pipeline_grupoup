import apache_beam as beam
from apache_beam.io import fileio
from apache_beam.options.pipeline_options import PipelineOptions


class read_sheets(beam.DoFn):
    def process(self, file_path):
        from openpyxl import load_workbook
        from apache_beam.io.gcp.gcsio import GcsIO

        gcs = GcsIO()
        with gcs.open(file_path) as f: 
            excel_file = load_workbook(f,read_only=True,data_only=True,keep_links=False)
            period = file_path[file_path.find("(")+1:file_path.find(")")]
            
            delete_sheets = ['83RINCONVESP','TOTAL TORTILLERIA ','ENTREGAS ABA','Hoja1','Hoja2','GENERAL ','REPARTO84','Hoja4','Hoja3','5','6']
            valid_sheets = list(set(excel_file.sheetnames) - set(delete_sheets))
            
            for sheet_name in valid_sheets:
                sheet = excel_file[sheet_name]                            
                columns = next(sheet.iter_rows(min_row=2,values_only=True))
                for row in sheet.iter_rows(min_row=3,values_only=True):
                    if row[0] is not None:
                        if len(columns) == len(row):
                            res = dict(zip(columns, row))
                            yield res, sheet_name, period
            excel_file.close()
            

class AddMetaData(beam.DoFn):
    def process(self, element):
        element[0]['sucursal']  =  element[1]
        element[0]['periodo']  =  element[2]
        yield element[0]

class clean_dict(beam.DoFn):
    def __init__(self,keep_list,rename):
        self.keep_list = keep_list
        self.rename = rename
        
    def process(self,element):
        from datetime import datetime

        new_dict = {}
        for key, value in element.items():
            if key in self.keep_list:
                if key == 'FECHA':
                    try:
                        new_dict[self.rename.get(key, key)] = datetime.strptime(str(value.day) + '-' + element['periodo'], '%d-%m-%y').date()
                    except (ValueError, KeyError,AttributeError):
                        new_dict[self.rename.get(key, key)] = None
                elif key == 'sucursal':
                    new_dict[self.rename.get(key, key)] = value
                else:
                    try:
                        new_dict[self.rename.get(key, key)] = float(value)
                    except (ValueError,TypeError):
                        new_dict[self.rename.get(key, key)] = 0
        yield new_dict
            
def run_pipeline():
    
    # Pipeline options
    beam_options = PipelineOptions(
    runner='DataflowRunner',
    project='proyectoanalisisgrupoup',
    temp_location='gs://ds-edw-raw-2b4dd4c8/tmp/',
    staging_location='gs://ds-edw-raw-2b4dd4c8/stg/',
    region ='us-central1',
    dataflow_service_options=['enable_prime'])
    
    with beam.Pipeline(options = beam_options) as pipeline:
        excel_files = (
            pipeline
            | 'List Excel Files' >> fileio.MatchFiles('gs://ds-edw-raw-2b4dd4c8/archivos_cierre/tortilla/*.xlsx')
            | 'Get metadata' >> fileio.ReadMatches()
            | 'Get Excel Paths' >> beam.Map(lambda file: file.metadata.path)
        )
        
        keep_list = {'FECHA','sucursal','INICIAL', 'ENTRADAS', 'ENTRADAS CAMION', 'SALIERON', 'ELABORO', 'EXISTENCIA ACTUAL', 'TOTAL TRASPASOS', 'REPARTO','TOTAL EN SACOS', 'DESPERDICIO TORTILLA', 'DESPERDICIO DE MASA', 'TORTILLA QUE QUEDO', 'TOTAL DESPERDICIO', 'TOTAL VENTA EN MOSTRADOR', 'TOTAL VENTA EN EFECTIVO', 'TOTAL VENTA CANASTA', 'TOTAL VENTA REFRESCO', 'TOTAL VENTA HECTOR', 'TOTAL VENTA GLOBAL', 'MEDIAS', 'VENTA DEL DIA', 'RAYA', 'VALES', 'MEDIAS DOS', 'DIF.DE PRECIOS', 'TOTAL CIERRE FLUJO'}
        rename = {'FECHA':'fecha','INICIAL': 'inicial', 'ENTRADAS': 'entradas', 'ENTRADAS CAMION': 'entradas_camion', 'SALIERON': 'salieron', 'ELABORO': 'elaboro', 'EXISTENCIA ACTUAL': 'existencia_actual','REPARTO':'total_reparto', 'TOTAL TRASPASOS': 'total_traspasos', 'TOTAL EN SACOS': 'total_en_sacos', 'DESPERDICIO TORTILLA': 'desperdicio_tortilla', 'DESPERDICIO DE MASA': 'desperdicio_de_masa', 'TORTILLA QUE QUEDO': 'tortilla_que_quedo', 'TOTAL DESPERDICIO': 'total_desperdicio', 'TOTAL VENTA EN MOSTRADOR': 'total_venta_en_mostrador', 'TOTAL VENTA EN EFECTIVO': 'total_venta_en_efectivo', 'TOTAL VENTA CANASTA': 'total_venta_canasta', 'TOTAL VENTA REFRESCO': 'total_venta_refresco', 'TOTAL VENTA HECTOR': 'total_venta_hector', 'TOTAL VENTA GLOBAL': 'total_venta_global', 'MEDIAS': 'medias', 'VENTA DEL DIA': 'venta_del_dia', 'RAYA': 'raya', 'VALES': 'vales', 'MEDIAS DOS': 'medias_dos', 'DIF.DE PRECIOS': 'dif_de_precios', 'TOTAL CIERRE FLUJO': 'total_cierre_flujo' }
        
        data = (
            excel_files
            | 'Read Excel' >>  beam.ParDo(read_sheets())
            | 'Add metadata' >> beam.ParDo(AddMetaData())
            | 'Clean format' >> beam.ParDo(clean_dict(keep_list,rename))
            | 'Filter errors' >> beam.Filter(lambda x: x['fecha'] is not None)
        )
        
                # Write to BigQuery
        export = (data | 'Write to BigQuery' >> beam.io.WriteToBigQuery(
            table='tortilla',
            dataset='bd_grupoup',
            project='proyectoanalisisgrupoup',
            create_disposition=beam.io.BigQueryDisposition.CREATE_NEVER,
            write_disposition=beam.io.BigQueryDisposition.WRITE_TRUNCATE,
            )
        )

if __name__ == '__main__':
    run_pipeline()